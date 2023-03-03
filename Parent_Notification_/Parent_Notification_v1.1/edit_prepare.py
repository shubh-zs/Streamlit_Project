from openpyxl import Workbook, load_workbook 
import os
a = True
def status(path_up_file):
    wb = load_workbook(path_up_file)
    sheet = wb.create_sheet('Parent')        
    sheet['A1'] = "Pylenin"
    sheet['B1'] = "loves"
    sheet['C1'] = "Python"
    wb.save(path_up_file)

    #Now check if the edit is done correctly by probing randomly and checking if they match with information we know!

    return False
