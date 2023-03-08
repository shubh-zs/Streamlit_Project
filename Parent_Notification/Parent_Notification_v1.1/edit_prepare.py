from openpyxl import Workbook, load_workbook 
import random

def status(atten_path,parent_dta,filt_val):      
    
    wb_attendance = load_workbook(atten_path)
    wb_parent = load_workbook(parent_dta)
    att = wb_attendance.active
    dta = wb_parent.active
    wb_attendance.create_sheet("Edit Data")
    
    filt_std = {}                                          #This Dictionary is used to store Roll number of Student, Then the Roll number and Parent contact info
    row_no = 1
    for at in att["B"]:
        if(row_no!=1): 
            if(at.value is not None and at.value<=filt_val):
                filt_std[att["A"+str(row_no)].value] = at.value
        row_no += 1
    parent_info = {}                                       #Is saving the names of parents in dictionary even necessary it can be done in a list
    row_no = 1
    for roll in dta["A"]:
        if(row_no!=1):
            if(roll.value in filt_std and roll.value is not None):
                filt_std[roll.value] = dta["C"+str(row_no)].value
                parent_info[roll.value] = dta["B"+str(row_no)].value
        row_no += 1 

    print(filt_std)
    print(parent_info)
    # wb_attendance.save(path_up_file)

    edit = wb_attendance["Edit Data"]
    row_no = 2
    for i in parent_info:                                   
        edit["A"+str(row_no)].value = i                     #Should I check if the roll number are in order present in both of the dictionary
        edit["B"+str(row_no)].value = parent_info[i]        #I think at this time i am not able to think of any edge case that will defy the assumption
        edit["C"+str(row_no)].value = filt_std[i]
        row_no += 1

    wb_attendance.save(atten_path) 

    #Now check if the edit is done correctly by probing randomly and checking if they match with information we know!

    probe1 = random.choice(list(filt_std.keys()))
    indx1 = list(filt_std.keys()).index(probe1)+2           

    probe2 = random.choice(list(filt_std.keys()))              #Assuming that probe values are unique
    indx2 = list(filt_std.keys()).index(probe2)+2
    
    # print(probe1,":",probe2)                                 #These lines of code were written to debug 
    # print(indx1,":",indx2)

    # print(
    #     edit["A"+str(indx1)].value,
    #     edit["B"+str(indx1)].value,
    #     edit["C"+str(indx1)].value
    # )

    def probe(indx,probe):
        if(edit["A"+str(indx)].value == probe):
            if(edit["B"+str(indx)].value == parent_info[probe]):
                if(edit["C"+str(indx)].value == filt_std[probe]):
                    return True
        return False
    
    if probe(indx1,probe1):
        if(probe(indx2,probe2)):
            return True
    
    #Assuming the editing was not successful
    # wb_attendance.remove_sheet("Edit Data")
    return False
